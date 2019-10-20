# -*- coding: utf-8 -*-
#import xlrd
#import xlwt
import sys
import os
import openpyxl
from openpyxl.styles import numbers

class excelWorker:
    # 初始化构造函数
    def __init__(self):
        # 多文件游标
        self.cursor = 1
        # 取出城市标签，如 beijing
        self.xlsPathIdentifier =  str(sys.argv[1])
        # 最终整表数据个数
        self.totalCount = 0
        # 获取 xls 文件集合
        self.xlsFiles = []
        path = os.path.join(os.getcwd(),self.xlsPathIdentifier)
        # 结果文件路径
        self.resultPath = os.path.join(os.getcwd(),self.xlsPathIdentifier,u'{0}全部数据.xlsx'.format(self.xlsPathIdentifier))
        # 先删除旧文件
        if os.path.exists(self.resultPath):
            os.remove(self.resultPath)
        
        for root, dirs, files in os.walk(path, topdown=False):
            for name in files:
                if '.xlsx' in name:
                    self.xlsFiles.append(os.path.join(root,name))
                    #inwb  = load_workbook(xlsFiles[0])
                    #获取第一个sheet内容
                    #ws = inwb.get_sheet_by_name(sheetnames[0])
                    #count = count + ws.max_row
        
    # 整合多个 xls 文件到一张 xlsx 表
    def combine(self):
        if len(self.xlsFiles) <= 0:
            print(u'没有找到任何文件')
            exit()
        # 创建文件
        outwb = openpyxl.Workbook()#打开一个将写的文件
        outws = outwb.create_sheet(title=u'在售列表')#在将写的文件创建sheet
        # style
        font = openpyxl.styles.Font(u'Arial', size = 10, color='000000')
        
        # 写入 sheet 头
        inwb = openpyxl.load_workbook(self.xlsFiles[0])
        inws = inwb[u'在售列表'] # OR inwb.sheetnames[1]，不知道为啥会自动创建一个名为 sheet 的默认表单
        cols = inws.max_column
        
        # 注释：openpyxl 与 xlrd/xlwt 的区别在于，前者的行列号从 1 开始，后者从 0 开始。
        for i in range(1,cols + 1):
            outws.cell(1,i).value = inws.cell(1,i).value
            outws.cell(1,i).font = font
        self.cursor = 2
        
        # 批量写入
        fileIndex = 1
        for xlsFile in self.xlsFiles:
            inwb = openpyxl.load_workbook(xlsFile)
            inws = inwb[u'在售列表']
            nrows = inws.max_row - 1 # 去掉标题
            ncols = inws.max_column
            
            print(u'{0} 数据条目：{1}'.format(xlsFile,nrows))
            for row in range(2,nrows+1+1):
                for col in range(1,ncols+1):
                    # 第一列是序号，做一个连贯的
                    if col == 1:
                        outws.cell(self.cursor,col).value = self.cursor - 1
                        outws.cell(self.cursor,col).number_format = '0' 
                    # 建成时间：把“未知”改成 -1
                    elif col == 10:
                        if inws.cell(row,col).value == u'未知':
                            outws.cell(self.cursor,col).value = -1
                        else:
                            outws.cell(self.cursor,col).value = int(inws.cell(row,col).value)
                        outws.cell(self.cursor,col).number_format = '0' 
                    # 房屋单价 房屋总价
                    elif col in [4,5,7]:
                        outws.cell(self.cursor,col).value = float(inws.cell(row,col).value)
                        outws.cell(self.cursor,col).number_format = '0.00'
                    # 建筑面积 关注人数
                    elif col in [7,33]:
                        outws.cell(self.cursor,col).value = float(inws.cell(row,col).value)
                        outws.cell(self.cursor,col).number_format = '0'
                    else:
                        outws.cell(self.cursor,col).value = inws.cell(row,col).value
                    outws.cell(self.cursor,col).font = font
                    
                self.cursor = self.cursor + 1
                self.totalCount = self.totalCount + 1
            # 每个文件保存一次，防止最后一次性写入的那种假死态
            #outwb.save(self.resultPath)
            #print(u'[{0}/{1}] 写入完毕'.format(fileIndex,len(self.xlsFiles)))
            #fileIndex = fileIndex + 1
        # 保存
        #resultXLS.save(self.resultPath)
        print(u'正在保存文件 {0} 请稍后'.format(self.xlsPathIdentifier))
        outwb.save(self.resultPath)
        print(u'整合最终数据条目 {0}'.format(self.totalCount))

worker = excelWorker()
worker.combine()