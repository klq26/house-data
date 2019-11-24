# -*- coding: utf-8 -*-
import xlrd
#import xlwt
import openpyxl


class generateExcel:
    def __init__(self):
        self.raws = []
        #self.wb = xlwt.Workbook(encoding='utf-8')
        # openxlpy 适配
        self.wb = openpyxl.Workbook()
        # style
        font = openpyxl.styles.Font(u'Arial', size=10, color='000000')
        self.font = font

    def writeExcelPositon(self, rowNumber, column, source_data):
        #self.ws.write(rowNumber, column, source_data)
        # openxlpy 适配
        self.ws.cell(rowNumber + 1, column + 1).value = source_data
        self.ws.cell(rowNumber + 1, column + 1).font = self.font

    def saveExcel(self, name):
        self.wb.save(name)

    def addSheetExcel(self, sheetName):
        #self.ws = self.wb.add_sheet(sheetName,cell_overwrite_ok=True)
        # openxlpy 适配
        self.ws = self.wb.active
        self.ws.title = sheetName
